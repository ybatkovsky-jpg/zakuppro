"""
Create test Excel file with Russian headers for BOM extraction testing.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from pathlib import Path

# Create output directory
output_dir = Path(__file__).parent
output_file = output_dir / "sample_bom.xlsx"

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BOM"

# Define headers with Russian column names
headers = ["Артикул", "Наименование", "Количество", "Поставщик"]

# Sample BOM data - realistic Russian invoice items
sample_data = [
    ["CPU-001", "Процессор Intel Core i7-12700K", 5, "Distributor LLC"],
    ["RAM-002", "Оперативная память DDR4 16GB", 10, "Memory Solutions"],
    ["SSD-003", "SSD 512GB SATA III", 8, "Storage Inc"],
    ["MB-004", "Материнская плата Z690", 3, "Components Pro"],
    ["GPU-005", "Видеокарта RTX 3060", 4, "Graphics Corp"],
    ["PSU-006", "Блок питания 650W", 6, "Power Systems"],
    ["CASE-007", "Корпус ATX Mid Tower", 2, "Case Manufacturer"],
    ["FAN-008", "Кулер CPU 120mm", 12, "Cooling Solutions"],
    # Empty row to test cleanup
    [],
    ["CABLE-009", "Кабель SATA 1m", 20, "Cables Inc"],
    ["MON-010", "Монитор 27 IPS", 1, "Display Tech"],
]

# Add header row with styling
header_font = Font(bold=True, size=12)
header_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = thin_border

# Add data rows
for row_idx, row_data in enumerate(sample_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        if row_data:  # Only border non-empty rows
            cell.border = thin_border

# Auto-adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column].width = adjusted_width

# Save the workbook
wb.save(output_file)
print(f"Test Excel file created: {output_file}")
print(f"Contains {len(sample_data)} rows including 1 empty row for cleanup testing")
